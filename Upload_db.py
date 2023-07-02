from sqlalchemy import create_engine, text
import openpyxl as xl

def db_connection():
  db_connection_string = "mysql+pymysql://rct26osfx6yxi9ml2tf6:pscale_pw_3GezCF2721gQEt3DZL9oYDrSgiXarePZDbLe5MyYrRN@aws.connect.psdb.cloud/inventory?charset=utf8mb4"
  # mecl_db_connection_string = "mysql+pymysql://mec_wh"
  engine = create_engine(
    db_connection_string, 
    connect_args={
      "ssl": {
        "ssl_ca": "/etc/ssl/cert.pem"
      }
    })
  return engine

def upload_file_to_db(file):

  engine = db_connection()

  # Open the workbook and define the worksheet
  # file = "selected_values"
  ext = ".xlsx"
  book = xl.load_workbook(file+ext, data_only=True)
  sheet = book['Sheet1']

  # Establish a MySQL connection
  database = engine.raw_connection()

  # Get the cursor, which is used to traverse the database, line by line
  cursor = database.cursor()

  # Create Table
  v1 = sheet.cell(1,1).value
  v2 = sheet.cell(1,2).value
  v3 = sheet.cell(1,3).value
  v4 = sheet.cell(1,4).value
  v5 = sheet.cell(1,5).value
  v6 = sheet.cell(1,6).value

  count_sheet_table = ("CREATE TABLE IF NOT EXISTS "+str(file)+"_master(id INT NOT NULL AUTO_INCREMENT,"+str(v1)+" LONGTEXT,"+str(v2)+" LONGTEXT,"+str(v3)+" LONGTEXT,"+str(v4)+" LONGTEXT,"+str(v5)+" LONGTEXT,"+str(v6)+" LONGTEXT,count LONGTEXT, variance LONGTEXT, count2 LONGTEXT, PRIMARY KEY (id) ) ")

  # Execute create table query
  cursor.execute(count_sheet_table)

  # Create the INSERT INTO SQL query
  query = "INSERT INTO "+str(file)+"_master ("+str(v1)+","+str(v2)+","+str(v3)+","+str(v4)+","+str(v5)+","+str(v6)+",count,variance,count2) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)"

  # Create a For loop to iterate through each row in the XLS file

  for r in range(2,sheet.max_row):
    site_name = sheet.cell(r,1).value

    catagory = sheet.cell(r,2).value

    article = sheet.cell(r,3).value
    
    sloc = sheet.cell(r,4).value
    
    value = sheet.cell(r,5).value
    
    qty = sheet.cell(r,6).value

    count = sheet.cell(r,7).value

    variance = sheet.cell(r,8).value

    count2 = sheet.cell(r,9).value

    # Assign values from each row        
    values = (site_name, catagory, article,sloc, value, qty, count, variance, count2)

    # Execute sql Query
    cursor.execute(query, values)

  # Close the cursor
  cursor.close()

  # Commit the transaction
  database.commit()

  # Close the database connection
  database.close()